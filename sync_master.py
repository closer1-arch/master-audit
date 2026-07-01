#!/usr/bin/env python3
"""
sync_master.py — ETL determinístico Sync/Backfill da Master de Auditoria Comercial.

Substitui o Apps Script "SYNC DE FECHAMENTOS v2" (quebrado). Lê as planilhas de
fechamento (.xlsx no Drive), faz upsert por Chave na master nativa, in-place.

NUNCA: deleta linha/aba, converte .xlsx em nativo, escreve colunas manuais,
toca formatação de linha existente. (Spec §9.)

Uso típico (2 passos):
    python sync_master.py            # DRY-RUN: só relata, não escreve (padrão)
    python sync_master.py --apply    # executa de verdade, após o Raul conferir

Cron contínuo (Railway, 2x/dia):
    python sync_master.py --mode ativo --apply

Credencial: service account JSON, via --sa ou env ALF_SYNC_SA_JSON
(ou GOOGLE_APPLICATION_CREDENTIALS). O valor pode ser um CAMINHO de arquivo
(dev local) OU o próprio CONTEÚDO JSON (Railway, que só tem env string). A
credencial NUNCA vai pro vault, log ou chat. A SA precisa: leitura nas fontes
(Drive) + edição na master (Sheets).
"""

from __future__ import annotations

import argparse
import io
import json
import os
import sys
import traceback
from datetime import datetime, timedelta, timezone

import sync_core as core

# --------------------------------------------------------------------------- #
# Alvo (§4).
# --------------------------------------------------------------------------- #
MASTER_ID = "1f2C3XvnHfYm2W9L0zrZ2ynhwXkJswB7XwOPRKOM5ES8"

# --------------------------------------------------------------------------- #
# Fontes (§3). O nome do closer é LITERAL e tem que bater com a master (a Chave
# depende disso). 'ativo' = sync contínuo; 'backfill' = one-time (idempotente).
# --------------------------------------------------------------------------- #
SOURCES = [
    {"closer": "João Guilherme",   "nicho": "Auxílio-Acidente",
     "file_id": "1-pyIcsRH18Q5WmSTFvL2H5MN75y8JFOv", "role": "ativo"},
    {"closer": "Beatriz de Souza", "nicho": "Auxílio-Acidente",
     "file_id": "1h0XxHSK_pm-Dto8K1gMCevGdobNGBY4n", "role": "ativo"},
    {"closer": "Maiara Mendes",    "nicho": "Vícios de Construção",
     "file_id": "17W2IULfceGtBuni1SikJfJeDArHmaiHz", "role": "ativo"},
    {"closer": "Raul Gabriel",     "nicho": "Vícios de Construção",
     "file_id": "1SVFv3eKXKltLB9a0nX4nG1pHA_Pjkj8K", "role": "backfill"},
    {"closer": "Catarina Prata",   "nicho": "Vícios de Construção",
     "file_id": "1Kg5FkBcLUXr1dFE9Nn0KcUIw6hSlIyal", "role": "backfill"},
]
# Fora do sync (§3): Priscila Gomes — Sheet nativa, dados já completos. Não ler.

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets",
]
GMT3 = timezone(timedelta(hours=-3))

DEFAULT_MAX_INSERTS = 60
ATIVOS = ("João Guilherme", "Beatriz de Souza", "Maiara Mendes")


def now_brt(fmt="%Y-%m-%d %H:%M"):
    return datetime.now(GMT3).strftime(fmt)


# --------------------------------------------------------------------------- #
# Credencial da service account: caminho de arquivo OU conteúdo JSON.
# resolve_sa_source é pura (sem Google) -> testável. get_services faz o I/O.
# --------------------------------------------------------------------------- #
def resolve_sa_source(sa_value):
    """
    Decide se sa_value é um caminho de arquivo existente (dev local) ou o próprio
    conteúdo JSON da SA (Railway). Devolve ('file', path) ou ('info', dict).
    Levanta ValueError se não for nenhum dos dois.
    """
    if sa_value and os.path.isfile(sa_value):
        return "file", sa_value
    try:
        info = json.loads(sa_value)
    except (TypeError, ValueError):
        info = None
    if isinstance(info, dict):
        return "info", info
    raise ValueError(
        "ALF_SYNC_SA_JSON/--sa não é caminho de arquivo válido nem JSON de "
        "service account."
    )


# --------------------------------------------------------------------------- #
# Serviços Google (import lazy: o sync_core e os testes não dependem disto).
# --------------------------------------------------------------------------- #
def get_services(sa_value):
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError:
        sys.exit("Faltam dependências. Rode: pip install -r requirements.txt")
    kind, val = resolve_sa_source(sa_value)
    if kind == "file":
        creds = service_account.Credentials.from_service_account_file(val, scopes=SCOPES)
    else:
        creds = service_account.Credentials.from_service_account_info(val, scopes=SCOPES)
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)
    sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)
    return drive, sheets


# --------------------------------------------------------------------------- #
# Notificação por e-mail (SMTP Gmail, STARTTLS). Best-effort: falha de SMTP não
# derruba o run. Só runs --apply notificam; dry-run não.
# --------------------------------------------------------------------------- #
def subject_ok(when):
    return f"✅ Sync master OK — {when} BRT"


def subject_abort(n_inserts, max_inserts):
    return f"🚨 Sync master ABORTADO — inserts={n_inserts} > teto {max_inserts}"


def subject_fail(when):
    return f"🚨 Sync master FALHOU — {when}"


def notify_email(subject, body):
    """Envia e-mail via smtp.gmail.com:587. Devolve True/False; nunca levanta."""
    user = os.environ.get("ALF_SYNC_SMTP_USER")
    pw = os.environ.get("ALF_SYNC_SMTP_PASS")
    to = os.environ.get("ALF_SYNC_MAIL_TO")
    if not (user and pw and to):
        print("notify_email: credenciais SMTP ausentes "
              "(ALF_SYNC_SMTP_USER/PASS/MAIL_TO); e-mail não enviado.")
        return False
    import smtplib
    from email.message import EmailMessage

    msg = EmailMessage()
    msg["From"] = user
    msg["To"] = to
    msg["Subject"] = subject
    msg.set_content(body)
    try:
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=30) as s:
            s.starttls()
            s.login(user, pw)
            s.send_message(msg)
        print(f"E-mail enviado: {subject}")
        return True
    except Exception as e:  # best-effort: reporta, não engole o resultado do run
        print(f"notify_email: falha no SMTP ({e}); run não afetado.")
        return False


def insert_guard_exceeded(n_inserts, max_inserts):
    """Guard de segurança: True quando o volume de inserts passa do teto."""
    return n_inserts > max_inserts


def download_xlsx(drive, file_id):
    from googleapiclient.http import MediaIoBaseDownload
    req = drive.files().get_media(fileId=file_id, supportsAllDrives=True)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------- #
# Leitura das fontes (.xlsx -> registros transformados).
# --------------------------------------------------------------------------- #
def read_source(drive, src, now_str):
    from openpyxl import load_workbook
    buf = download_xlsx(drive, src["file_id"])
    wb = load_workbook(buf, read_only=True, data_only=True)
    rows = []
    skipped_tabs = []
    try:
        for ws in wb.worksheets:
            block = core.parse_block(ws.title)
            if block is None:
                skipped_tabs.append(ws.title)
                continue
            bmonth, byear = block
            for raw in ws.iter_rows(values_only=True):
                rec = core.transform_row(list(raw), src["closer"], src["nicho"],
                                         bmonth, byear, now_str)
                if rec:
                    rows.append(rec)
    finally:
        wb.close()
    return rows, skipped_tabs


# --------------------------------------------------------------------------- #
# Leitura da master (mapa Chave -> posição, último dado por aba).
# --------------------------------------------------------------------------- #
def read_master(sheets, master_id):
    meta = sheets.spreadsheets().get(spreadsheetId=master_id).execute()
    props = {s["properties"]["title"]: s["properties"] for s in meta["sheets"]}

    key_index = {}     # chave -> (tab, rownum 1-based)
    tab_lastrow = {}   # tab -> última linha com dado (1-based; 1 = só cabeçalho)
    for title in props:
        rng = f"'{title}'!A2:O"
        vals = (sheets.spreadsheets().values()
                .get(spreadsheetId=master_id, range=rng).execute().get("values", []))
        tab_lastrow[title] = 1 + len(vals)
        for i, row in enumerate(vals):
            chave = row[core.COL_CHAVE] if len(row) > core.COL_CHAVE else ""
            if chave:
                key_index[chave] = (title, i + 2)
    return props, key_index, tab_lastrow


# --------------------------------------------------------------------------- #
# Criação de aba nova (cabeçalho + formatação + Chave oculta).
# --------------------------------------------------------------------------- #
def create_tab(sheets, master_id, title):
    resp = sheets.spreadsheets().batchUpdate(
        spreadsheetId=master_id,
        body={"requests": [{"addSheet": {"properties": {
            "title": title,
            "gridProperties": {"frozenRowCount": 1},
        }}}]},
    ).execute()
    sheet_id = resp["replies"][0]["addSheet"]["properties"]["sheetId"]

    sheets.spreadsheets().values().update(
        spreadsheetId=master_id, range=f"'{title}'!A1:O1",
        valueInputOption="RAW", body={"values": [core.HEADER]},
    ).execute()

    sheets.spreadsheets().batchUpdate(spreadsheetId=master_id, body={"requests": [
        {"repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1},
            "cell": {"userEnteredFormat": {
                "textFormat": {"bold": True},
                "backgroundColor": {"red": 0.85, "green": 0.85, "blue": 0.85},
            }},
            "fields": "userEnteredFormat(textFormat,backgroundColor)",
        }},
        {"updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "COLUMNS",
                      "startIndex": core.COL_CHAVE, "endIndex": core.COL_CHAVE + 1},
            "properties": {"hiddenByUser": True}, "fields": "hiddenByUser",
        }},
    ]}).execute()
    return sheet_id


# --------------------------------------------------------------------------- #
# Aplicação do plano (updates + inserts).
# --------------------------------------------------------------------------- #
def apply_updates(sheets, master_id, updates, key_index):
    data = []
    for rec in updates:
        tab, rownum = key_index[rec["chave"]]
        data.extend(core.update_segments(tab, rownum, rec))
    for i in range(0, len(data), 500):  # lotes para não estourar a request
        chunk = data[i:i + 500]
        sheets.spreadsheets().values().batchUpdate(
            spreadsheetId=master_id,
            body={"valueInputOption": "USER_ENTERED",
                  "data": [{"range": r, "values": v} for r, v in chunk]},
        ).execute()


def apply_inserts(sheets, master_id, inserts, props, tab_lastrow):
    by_tab = {}
    for rec in inserts:
        by_tab.setdefault(rec["target_tab"], []).append(rec)

    for tab, recs in sorted(by_tab.items()):
        if tab not in props:
            create_tab(sheets, master_id, tab)
            props[tab] = {"title": tab}
            tab_lastrow[tab] = 1
        start = tab_lastrow[tab] + 1
        block = [core.insert_row_values(r) for r in recs]
        end = start + len(block) - 1
        sheets.spreadsheets().values().update(
            spreadsheetId=master_id, range=f"'{tab}'!A{start}:O{end}",
            valueInputOption="USER_ENTERED", body={"values": block},
        ).execute()
        tab_lastrow[tab] = end


# --------------------------------------------------------------------------- #
# Relatório (dry-run e pós-run). Devolve o texto (reusado no corpo do e-mail).
# --------------------------------------------------------------------------- #
def format_report(all_rows, inserts, updates, props, source_stats):
    out = []
    out.append("=" * 64)
    out.append("RESUMO DO SYNC DA MASTER DE AUDITORIA COMERCIAL")
    out.append("=" * 64)

    out.append("\nPor closer (linhas válidas lidas / reparadas / sem-data):")
    for closer, st in source_stats.items():
        out.append(f"  {closer:<18} lidas={st['rows']:>4}  "
                   f"reparadas={st['repaired']:>3}  sem-data={st['sd']:>2}")
        if st["skipped_tabs"]:
            out.append(f"      abas ignoradas: {', '.join(st['skipped_tabs'])}")

    ins_by_tab, upd_by_tab = {}, {}
    for r in inserts:
        ins_by_tab[r["target_tab"]] = ins_by_tab.get(r["target_tab"], 0) + 1
    for r in updates:
        upd_by_tab[r["target_tab"]] = upd_by_tab.get(r["target_tab"], 0) + 1

    out.append("\nPor aba da master (insert / update | * = aba será criada):")
    for tab in sorted(set(ins_by_tab) | set(upd_by_tab)):
        flag = "" if tab in props else "  *NOVA*"
        out.append(f"  {tab:<16} insert={ins_by_tab.get(tab, 0):>4}  "
                   f"update={upd_by_tab.get(tab, 0):>4}{flag}")

    repaired = [r for r in all_rows if r["date_status"] in ("reconciled", "reformatted")]
    sd = [r for r in all_rows if r["date_status"] == "sd"]

    if repaired:
        out.append(f"\nDatas reparadas ({len(repaired)}):")
        for r in repaired:
            out.append(f"  [{r['date_status']:<11}] {r['closer']:<16} "
                       f"{r['nome'][:32]:<32} -> {r['data_iso']}")
    if sd:
        out.append(f"\nSem-data ({len(sd)}) — vão para a aba do bloco de origem:")
        for r in sd:
            out.append(f"  {r['closer']:<16} {r['nome'][:32]:<32} -> {r['target_tab']}")

    out.append("\n" + "-" * 64)
    out.append(f"TOTAL: {len(inserts)} inserts, {len(updates)} updates, "
               f"{len(repaired)} datas reparadas, {len(sd)} sem-data.")
    out.append("-" * 64)
    return "\n".join(out)


# --------------------------------------------------------------------------- #
# Núcleo do run (levanta exceção em falha — quem trata é o main).
# --------------------------------------------------------------------------- #
def _run(args):
    sources = SOURCES
    if args.mode != "all":
        sources = [s for s in sources if s["role"] == args.mode]
    if args.only:
        sources = [s for s in sources if s["closer"] in args.only]
    if not sources:
        sys.exit("Nenhuma fonte selecionada pelos filtros.")

    now_str = datetime.now(GMT3).strftime("%Y-%m-%d %H:%M:%S")
    mode_label = "APPLY (escrevendo)" if args.apply else "DRY-RUN (sem escrita)"
    print(f"Modo: {mode_label} | fontes: {', '.join(s['closer'] for s in sources)}")

    drive, sheets = get_services(args.sa)

    all_rows, source_stats = [], {}
    for src in sources:
        rows, skipped = read_source(drive, src, now_str)
        all_rows.extend(rows)
        source_stats[src["closer"]] = {
            "rows": len(rows),
            "repaired": sum(1 for r in rows if r["date_status"] in ("reconciled", "reformatted")),
            "sd": sum(1 for r in rows if r["date_status"] == "sd"),
            "skipped_tabs": skipped,
        }

    props, key_index, tab_lastrow = read_master(sheets, args.master_id)
    inserts, updates = core.build_plan(all_rows, set(key_index.keys()))

    report_text = format_report(all_rows, inserts, updates, props, source_stats)
    print("\n" + report_text + "\n")

    if not args.apply:
        print("DRY-RUN: nada foi escrito. Revise o resumo e rode com --apply para aplicar.")
        return

    # Guard de insert: volume anômalo aborta ANTES de qualquer escrita.
    max_inserts = int(os.environ.get("ALF_SYNC_MAX_INSERTS", str(DEFAULT_MAX_INSERTS)))
    if insert_guard_exceeded(len(inserts), max_inserts):
        reason = (f"Guard de insert acionado: {len(inserts)} inserts > teto "
                  f"{max_inserts}. NADA foi escrito (nem update nem insert). "
                  f"Ajuste ALF_SYNC_MAX_INSERTS ou investigue a origem.")
        print(reason)
        notify_email(subject_abort(len(inserts), max_inserts),
                     reason + "\n\n" + report_text)
        sys.exit(1)

    print("Aplicando updates...")
    apply_updates(sheets, args.master_id, updates, key_index)
    print("Aplicando inserts...")
    apply_inserts(sheets, args.master_id, inserts, props, tab_lastrow)
    print("Concluído. Master atualizada in-place.")
    notify_email(subject_ok(now_brt()), report_text)


# --------------------------------------------------------------------------- #
# Main: argparse + --test-email + try/except que notifica falha (só em --apply).
# --------------------------------------------------------------------------- #
def main(argv=None):
    ap = argparse.ArgumentParser(description="Sync/Backfill da Master de Auditoria Comercial.")
    ap.add_argument("--apply", action="store_true",
                    help="Executa de verdade (escreve na master). Sem isto = dry-run.")
    ap.add_argument("--sa", default=os.environ.get("ALF_SYNC_SA_JSON")
                    or os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"),
                    help="Caminho do SA JSON OU o próprio conteúdo JSON.")
    ap.add_argument("--mode", choices=["all", "ativo", "backfill"], default="all",
                    help="Quais fontes processar (default: all). 'ativo' = só os 3 "
                         "closers ativos (exclui backfill Raul/Catarina).")
    ap.add_argument("--only", action="append", default=[],
                    help="Filtra por closer (pode repetir). Ex.: --only 'Raul Gabriel'.")
    ap.add_argument("--master-id", default=MASTER_ID, help="Override do fileId da master.")
    ap.add_argument("--test-email", action="store_true",
                    help="Manda um e-mail de teste e sai (valida SMTP no Railway).")
    args = ap.parse_args(argv)

    if args.test_email:
        ok = notify_email("✅ Sync master — e-mail de teste",
                          f"Teste de credencial SMTP do cron ({now_brt()} BRT). "
                          "Se você recebeu isto, o envio funciona.")
        sys.exit(0 if ok else 1)

    if not args.sa:
        sys.exit("Service account não informado. Use --sa PATH ou env ALF_SYNC_SA_JSON.")

    try:
        _run(args)
    except SystemExit:
        raise  # aborts/erros de validação já reportados; não vira "FALHOU"
    except Exception:
        tb = traceback.format_exc()
        print(tb, file=sys.stderr)
        if args.apply:
            notify_email(subject_fail(now_brt()),
                         "Exceção não tratada no run --apply:\n\n" + tb)
        sys.exit(1)


if __name__ == "__main__":
    main()
